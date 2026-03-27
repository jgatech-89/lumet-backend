"""
Importación masiva de clientes desde Excel: validación previa (todo el archivo) y persistencia atómica.
"""
import unicodedata
from collections import defaultdict
from django.db import transaction, IntegrityError
from django.db.models import Q
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError as DRFValidationError

from apps.core.choices import TIPO_IDENTIFICACION
from apps.formularios.models import Campo as CampoForm
from apps.formularios.services import get_campos_formulario
from apps.servicio.models import Servicio as ServicioModel

from .models import FormularioCliente
from .serializers import ClienteCreateSerializer


class ExcelImportPersistError(Exception):
    """Errores de validación en la fase de persistencia (transacción revertida)."""

    def __init__(self, errores_items):
        self.errores_items = errores_items
        super().__init__()


# Nombres de columna idénticos a la fila de encabezados del Excel de importación
EXCEL_COLUMN_BY_SERIALIZER_FIELD = {
    'nombre': 'Nombre completo',
    'tipo_identificacion': 'Tipo identificación',
    'numero_identificacion': 'Número identificación',
    'cuenta_bancaria': 'Cuenta bancaria',
    'direccion': 'Dirección',
    'telefono': 'Teléfono',
    'correo_electronico_o_carta': 'Correo o carta o papel',
    'compania_anterior': 'Compañía anterior',
    'compania_actual': 'Compañía actual',
    'servicio_id': 'Compañía actual',
    'producto': 'Producto',
    'respuestas': 'Tipo cliente',
}


def _infer_columna_excel_desde_mensaje(msg):
    """Para non_field_errors u otras claves sin mapeo directo: deduce la columna por el texto del error."""
    m = (msg or '').lower()
    if 'número de identificación' in m or 'numero de identificacion' in m:
        return 'Número identificación'
    if 'cuenta bancaria' in m:
        return 'Cuenta bancaria'
    if 'correo' in m or 'carta' in m or 'papel' in m:
        return 'Correo o carta o papel'
    if 'servicio' in m and 'id' in m:
        return 'Compañía actual'
    if 'respuestas' in m and ('campo' in m or 'configurado' in m):
        return 'Tipo cliente'
    if 'cups' in m:
        if 'luz' in m:
            return 'CUPS LUZ'
        if 'gas' in m:
            return 'CUPS GAS'
        return 'CUPS LUZ'
    return 'Datos generales'


def _flatten_serializer_error_values(val):
    """Convierte valores anidados de serializer.errors en lista de mensajes str."""
    if val is None:
        return []
    if isinstance(val, dict):
        msgs = []
        for v in val.values():
            msgs.extend(_flatten_serializer_error_values(v))
        return msgs
    if isinstance(val, (list, tuple)):
        return [str(x) for x in val]
    return [str(val)]


def _append_errores_serializer_excel(serializer_errors, row_idx, errores):
    """Un error por campo, con columna = encabezado del Excel (no 'Datos generales' salvo último recurso)."""
    for key, val in serializer_errors.items():
        messages = _flatten_serializer_error_values(val)
        if not messages:
            messages = ['Error de validación']
        for msg in messages:
            if key == 'non_field_errors':
                col = _infer_columna_excel_desde_mensaje(msg)
            else:
                col = EXCEL_COLUMN_BY_SERIALIZER_FIELD.get(key)
                if not col:
                    col = _infer_columna_excel_desde_mensaje(msg)
            errores.append(_err(row_idx, col, msg))


def _norm_key(valor):
    txt = unicodedata.normalize('NFKD', str(valor or '')).encode('ascii', 'ignore').decode('ascii')
    return ' '.join(txt.strip().lower().split())


def _to_str(valor):
    if valor is None:
        return ''
    return str(valor).strip()


def _map_select(valor_excel, opciones_validas):
    if not _to_str(valor_excel):
        return ''
    mapa = {_norm_key(v): v for v in opciones_validas if _to_str(v)}
    return mapa.get(_norm_key(valor_excel))


def _err(row_idx, col_name, msg):
    return {
        'fila': row_idx,
        'columna': col_name,
        'mensaje': msg,
        'texto': f'Error en la fila {row_idx}, columna "{col_name}": {msg}',
    }


def _build_import_context():
    """Carga opciones y catálogos una sola vez por importación."""
    nombres_fibra = ['fibra', 'Fibra']
    q_fibra = Q()
    for n in nombres_fibra:
        q_fibra |= Q(nombre__iexact=n)
    campo_fibra = CampoForm.objects.filter(
        fecha_elimina__isnull=True, tipo='select'
    ).filter(q_fibra).prefetch_related('opciones').first()
    fibra_opciones_validas = set()
    nombre_campo_fibra = ''
    if campo_fibra:
        nombre_campo_fibra = campo_fibra.nombre
        for op in campo_fibra.opciones.filter(activo=True, estado='1'):
            v = (op.value or op.label or '').strip()
            if v:
                fibra_opciones_validas.add(v)

    tipos_identificacion = [v[0] for v in TIPO_IDENTIFICACION]
    servicios_activos = list(
        ServicioModel.objects.filter(estado='1', fecha_elimina__isnull=True).order_by('nombre')
    )
    servicios_por_nombre_norm = {_norm_key(s.nombre): s for s in servicios_activos if _to_str(s.nombre)}

    nombres_producto = ['producto', 'Productos', 'Tipo producto', 'tipo de producto']
    q_producto = Q()
    for n in nombres_producto:
        q_producto |= Q(nombre__iexact=n)
    campos_producto = CampoForm.objects.filter(
        fecha_elimina__isnull=True, tipo='select'
    ).filter(q_producto).prefetch_related('opciones')
    producto_opciones = []
    for cp in campos_producto:
        for op in cp.opciones.filter(activo=True, estado='1').order_by('orden', 'id'):
            v = _to_str(op.value or op.label)
            if v and v not in producto_opciones:
                producto_opciones.append(v)

    nombres_tipo_cliente = [
        'tipo_cliente', 'tipo de cliente', 'Tipo de cliente', 'Tipo Cliente', 'tipo cliente', 'Tipo cliente',
    ]
    q_tipo_cliente = Q()
    for n in nombres_tipo_cliente:
        q_tipo_cliente |= Q(nombre__iexact=n)
    campos_tipo_cliente = CampoForm.objects.filter(
        fecha_elimina__isnull=True, tipo='select'
    ).filter(q_tipo_cliente).prefetch_related('opciones')
    tipo_cliente_opciones = []
    for ctc in campos_tipo_cliente:
        for op in ctc.opciones.filter(activo=True, estado='1').order_by('orden', 'id'):
            v = _to_str(op.value or op.label)
            if v and v not in tipo_cliente_opciones:
                tipo_cliente_opciones.append(v)

    mantenimiento_opciones = ['si', 'no']

    return {
        'fibra_opciones_validas': fibra_opciones_validas,
        'nombre_campo_fibra': nombre_campo_fibra,
        'tipos_identificacion': tipos_identificacion,
        'servicios_por_nombre_norm': servicios_por_nombre_norm,
        'producto_opciones': producto_opciones,
        'tipo_cliente_opciones': tipo_cliente_opciones,
        'campos_tipo_cliente': campos_tipo_cliente,
        'mantenimiento_opciones': mantenimiento_opciones,
    }


def _resolver_nombre_campo_tipo_cliente(servicio_id, producto, servicio_obj, campos_tipo_cliente_qs):
    nombre_campo_tipo = 'Tipo cliente'
    if servicio_id and servicio_obj:
        norm_tc = lambda s: (s or '').lower().replace(' ', '_')
        patrones_tc = [
            'tipo_cliente', 'tipo de cliente', 'Tipo de cliente', 'Tipo Cliente', 'tipo cliente', 'Tipo cliente',
        ]
        for c in get_campos_formulario(servicio_obj.empresa_id, servicio_id, producto or None):
            if any(norm_tc(c.nombre) == norm_tc(p) for p in patrones_tc):
                return c.nombre
    elif campos_tipo_cliente_qs.exists():
        return campos_tipo_cliente_qs.first().nombre
    return nombre_campo_tipo


def _build_duplicate_maps(rows_list):
    """Índices de fila por número de identificación y cuenta (normalizados), solo filas no vacías."""
    numero_filas = defaultdict(list)
    cuenta_filas = defaultdict(list)
    for row_idx, row in rows_list:
        num = _to_str(row[2]) if len(row) > 2 else ''
        if num:
            numero_filas[num.strip().lower()].append(row_idx)
        cuenta = _to_str(row[3]) if len(row) > 3 else ''
        if cuenta:
            cuenta_filas[cuenta.strip().lower()].append(row_idx)
    return numero_filas, cuenta_filas


def _validate_row(row_idx, row, ctx, request, numero_filas, cuenta_filas):
    """
    Valida una fila sin persistir. Devuelve (spec, errores).
    spec es None si hay errores; si no, dict listo para persistir.
    """
    errores = []

    nombre = _to_str(row[0]) if len(row) > 0 else ''
    if not nombre:
        errores.append(_err(row_idx, 'Nombre completo', 'valor vacío'))

    tipo_id_raw = _to_str(row[1]) if len(row) > 1 else ''
    tipo_id = _map_select(tipo_id_raw, ctx['tipos_identificacion'])
    if tipo_id_raw and not tipo_id:
        errores.append(_err(row_idx, 'Tipo identificación', f'valor no válido: "{tipo_id_raw}"'))

    numero_id = _to_str(row[2]) if len(row) > 2 else ''
    cuenta_bancaria = _to_str(row[3]) if len(row) > 3 else ''
    direccion = _to_str(row[4]) if len(row) > 4 else ''
    telefono = _to_str(row[5]) if len(row) > 5 else ''
    correo_electronico_o_carta = _to_str(row[6]) if len(row) > 6 else ''
    compania_ant = _to_str(row[7]) if len(row) > 7 else ''
    compania_act_raw = _to_str(row[8]) if len(row) > 8 else ''
    producto_raw = _to_str(row[9]) if len(row) > 9 else ''
    cups_luz = _to_str(row[10]) if len(row) > 10 else ''
    cups_gas = _to_str(row[11]) if len(row) > 11 else ''
    tipo_cliente_raw = _to_str(row[12]) if len(row) > 12 else ''
    mantenimiento_raw = _to_str(row[13]) if len(row) > 13 else ''
    fibra_raw = _to_str(row[14]) if len(row) > 14 else ''

    servicio = None
    compania_act = ''
    if compania_act_raw:
        servicio = ctx['servicios_por_nombre_norm'].get(_norm_key(compania_act_raw))
        if not servicio:
            errores.append(_err(row_idx, 'Compañía actual', f'valor no válido: "{compania_act_raw}"'))
        else:
            compania_act = servicio.nombre or ''

    producto = ''
    if producto_raw:
        producto = _map_select(producto_raw, ctx['producto_opciones'])
        if not producto:
            errores.append(_err(row_idx, 'Producto', f'valor no válido: "{producto_raw}"'))

    mantenimiento = ''
    if mantenimiento_raw:
        mantenimiento = _map_select(mantenimiento_raw, ctx['mantenimiento_opciones'])
        if not mantenimiento:
            errores.append(_err(row_idx, 'Mantenimiento', f'valor no válido: "{mantenimiento_raw}"'))

    fibra = ''
    if fibra_raw:
        fibra = _map_select(fibra_raw, ctx['fibra_opciones_validas'])
        if not fibra:
            errores.append(_err(row_idx, 'Fibra', f'valor no válido: "{fibra_raw}"'))

    tipo_cliente = ''
    if tipo_cliente_raw:
        tipo_cliente = _map_select(tipo_cliente_raw, ctx['tipo_cliente_opciones'])
        if not tipo_cliente:
            errores.append(_err(row_idx, 'Tipo cliente', f'valor no válido: "{tipo_cliente_raw}"'))

    producto_norm = _norm_key(producto)
    es_luz = producto_norm == 'luz'
    es_gas = producto_norm == 'gas'
    es_luz_gas = producto_norm in ('luz gas', 'luz y gas', 'gas y luz')

    if es_luz and cups_gas:
        errores.append(_err(row_idx, 'CUPS GAS', 'no aplica para producto LUZ'))
    elif es_gas and cups_luz:
        errores.append(_err(row_idx, 'CUPS LUZ', 'no aplica para producto GAS'))
    elif not es_luz and not es_gas and not es_luz_gas:
        if cups_luz:
            errores.append(_err(row_idx, 'CUPS LUZ', f'no aplica para producto "{producto or "vacío"}"'))
        if cups_gas:
            errores.append(_err(row_idx, 'CUPS GAS', f'no aplica para producto "{producto or "vacío"}"'))

    servicio_id = servicio.id if servicio else None
    srv_tc = servicio
    if servicio_id and not srv_tc:
        try:
            srv_tc = ServicioModel.objects.get(pk=servicio_id, estado='1', fecha_elimina__isnull=True)
        except ServicioModel.DoesNotExist:
            srv_tc = None

    respuestas_import = []
    if tipo_cliente:
        nombre_campo_tipo = _resolver_nombre_campo_tipo_cliente(
            servicio_id, producto, srv_tc, ctx['campos_tipo_cliente'],
        )
        respuestas_import.append({
            'nombre_campo': nombre_campo_tipo,
            'respuesta_campo': tipo_cliente,
        })

    payload = {
        'servicio_id': servicio_id,
        'nombre': nombre,
        'tipo_identificacion': tipo_id,
        'numero_identificacion': numero_id,
        'telefono': telefono,
        'correo_electronico_o_carta': correo_electronico_o_carta,
        'direccion': direccion,
        'cuenta_bancaria': cuenta_bancaria,
        'compania_anterior': compania_ant,
        'compania_actual': compania_act,
        'producto': producto,
        'respuestas': respuestas_import,
    }

    serializer = ClienteCreateSerializer(data=payload, context={'request': request, 'importar_excel': True})
    if not serializer.is_valid():
        _append_errores_serializer_excel(serializer.errors, row_idx, errores)

    if numero_id:
        k = numero_id.strip().lower()
        filas_mismo_num = numero_filas.get(k) or []
        if len(filas_mismo_num) > 1:
            todas = ', '.join(str(f) for f in sorted(filas_mismo_num))
            errores.append(_err(
                row_idx, 'Número identificación',
                f'valor duplicado en el archivo (filas: {todas}).',
            ))

    if cuenta_bancaria:
        kc = cuenta_bancaria.strip().lower()
        filas_misma_cuenta = cuenta_filas.get(kc) or []
        if len(filas_misma_cuenta) > 1:
            todas = ', '.join(str(f) for f in sorted(filas_misma_cuenta))
            errores.append(_err(
                row_idx, 'Cuenta bancaria',
                f'valor duplicado en el archivo (filas: {todas}).',
            ))

    if errores:
        return None, errores

    spec = {
        'row_idx': row_idx,
        'payload': payload,
        'cups_luz': cups_luz,
        'cups_gas': cups_gas,
        'mantenimiento': mantenimiento,
        'fibra': fibra,
    }
    return spec, []


def _persist_row_after_create(cliente, spec, ctx):
    """Tras ClienteCreateSerializer.save(): CUPS, mantenimiento, fibra en FormularioCliente."""
    cliente.creado_por_carga_masiva = True
    cliente.save(update_fields=['creado_por_carga_masiva'])

    cups_luz = spec['cups_luz']
    cups_gas = spec['cups_gas']
    mantenimiento = spec['mantenimiento']
    fibra = spec['fibra']

    ce = cliente.cliente_empresas.filter(estado='1').order_by('id').first()
    nombre_campo_fibra = ctx['nombre_campo_fibra']
    fibra_opciones_validas = ctx['fibra_opciones_validas']

    if ce:
        campos_producto_cfg = list(get_campos_formulario(
            empresa_id=ce.empresa_id,
            servicio_id=ce.servicio_id,
            producto=ce.producto,
        ))
        cups_luz_nombre = next(
            (c.nombre for c in campos_producto_cfg
             if 'cups' in _norm_key(c.nombre) and 'luz' in _norm_key(c.nombre)),
            'CUPS LUZ',
        )
        cups_gas_nombre = next(
            (c.nombre for c in campos_producto_cfg
             if 'cups' in _norm_key(c.nombre) and 'gas' in _norm_key(c.nombre)),
            'CUPS GAS',
        )
    else:
        cups_luz_nombre = 'CUPS LUZ'
        cups_gas_nombre = 'CUPS GAS'

    if cups_luz and ce:
        FormularioCliente.objects.update_or_create(
            cliente=cliente,
            cliente_empresa=ce,
            nombre_campo=cups_luz_nombre,
            defaults={'respuesta_campo': str(cups_luz), 'estado': '1'},
        )
    if cups_gas and ce:
        FormularioCliente.objects.update_or_create(
            cliente=cliente,
            cliente_empresa=ce,
            nombre_campo=cups_gas_nombre,
            defaults={'respuesta_campo': str(cups_gas), 'estado': '1'},
        )
    if mantenimiento and mantenimiento in ('si', 'no') and ce:
        FormularioCliente.objects.update_or_create(
            cliente=cliente,
            cliente_empresa=ce,
            nombre_campo='Mantenimiento',
            defaults={'respuesta_campo': mantenimiento, 'estado': '1'},
        )
    if fibra and nombre_campo_fibra and fibra in fibra_opciones_validas and ce:
        FormularioCliente.objects.update_or_create(
            cliente=cliente,
            cliente_empresa=ce,
            nombre_campo=nombre_campo_fibra,
            defaults={'respuesta_campo': str(fibra), 'estado': '1'},
        )


def run_excel_import(request, archivo):
    """
    Fase 1: validar todas las filas.
    Fase 2: si no hay errores, persistir en una sola transacción atómica.
    """
    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        msg = str(e)
        err_item = {'fila': None, 'columna': 'Archivo', 'mensaje': msg, 'texto': msg}
        return {
            'http_status': 400,
            'body': {'success': False, 'errors': [err_item], 'errores': [msg]},
        }

    ctx = _build_import_context()
    rows_list = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        rows_list.append((row_idx, row))

    numero_filas, cuenta_filas = _build_duplicate_maps(rows_list)

    all_errors = []
    specs = []

    for row_idx, row in rows_list:
        spec, row_errors = _validate_row(row_idx, row, ctx, request, numero_filas, cuenta_filas)
        all_errors.extend(row_errors)
        if spec:
            specs.append(spec)

    wb.close()

    if all_errors:
        return {
            'http_status': 400,
            'body': {
                'success': False,
                'errors': all_errors,
                'errores': [e['texto'] for e in all_errors],
            },
        }

    def _flatten_drf_errors(detail):
        if detail is None:
            return 'Error de validación'
        if isinstance(detail, (list, tuple)):
            return '; '.join(_flatten_drf_errors(x) for x in detail)
        if isinstance(detail, dict):
            return '; '.join(f'{k}: {_flatten_drf_errors(v)}' for k, v in detail.items())
        return str(detail)

    creados = 0
    try:
        with transaction.atomic():
            for spec in specs:
                row_idx = spec.get('row_idx')
                try:
                    ser = ClienteCreateSerializer(
                        data=spec['payload'],
                        context={'request': request, 'importar_excel': True},
                    )
                    ser.is_valid(raise_exception=True)
                    cliente = ser.save()
                    _persist_row_after_create(cliente, spec, ctx)
                    creados += 1
                except DRFValidationError as e:
                    detalle = getattr(e, 'detail', e)
                    fila_errores = []
                    if isinstance(detalle, dict):
                        _append_errores_serializer_excel(detalle, row_idx, fila_errores)
                    elif isinstance(detalle, (list, tuple)):
                        for item in detalle:
                            msg = str(item)
                            fila_errores.append(
                                _err(row_idx, _infer_columna_excel_desde_mensaje(msg), msg)
                            )
                    if not fila_errores:
                        msg = _flatten_drf_errors(detalle)
                        fila_errores = [_err(row_idx, _infer_columna_excel_desde_mensaje(msg), msg)]
                    transaction.set_rollback(True)
                    raise ExcelImportPersistError(fila_errores) from e
    except ExcelImportPersistError as e:
        items = e.errores_items
        return {
            'http_status': 400,
            'body': {
                'success': False,
                'errors': items,
                'errores': [x['texto'] for x in items],
            },
        }
    except IntegrityError as e:
        msg = str(e)
        err_item = {'fila': None, 'columna': 'Base de datos', 'mensaje': msg, 'texto': msg}
        return {
            'http_status': 400,
            'body': {
                'success': False,
                'errors': [err_item],
                'errores': [msg],
            },
        }
    except Exception as e:
        msg = str(e)
        err_item = {'fila': None, 'columna': 'Sistema', 'mensaje': msg, 'texto': msg}
        return {
            'http_status': 500,
            'body': {
                'success': False,
                'errors': [err_item],
                'errores': [msg],
            },
        }

    return {
        'http_status': 200,
        'body': {
            'success': True,
            'creados': creados,
            'mensaje': f'Se importaron {creados} cliente(s) correctamente.',
        },
    }
