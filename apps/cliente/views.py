import io
from django.http import HttpResponse
from django.db.models import Prefetch
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.filters import SearchFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

from .models import Cliente, FormularioCliente, HistorialEstadoVenta, ClienteEmpresa
from .serializers import (
    ClienteSerializer,
    ClienteCreateSerializer,
    ClienteDetalleSerializer,
    ClienteUpdateSerializer,
    ClienteAgregarProductoSerializer,
    ClienteActualizarProductoSerializer,
    _cambiar_estado_venta,
)
from .filters import ClienteFilter


def _estado_venta_cliente(cliente):
    """Estado a nivel cliente: prefiere historial legacy (sin producto); si no, el primer activo."""
    # Preferir legacy (cliente_empresa null) para no tomar el del último producto
    h_legacy = next(
        (x for x in cliente.historial_estados_venta.all() if x.cliente_empresa_id is None and x.activo),
        None,
    )
    if h_legacy:
        return h_legacy.estado or 'venta_iniciada'
    h = next((x for x in cliente.historial_estados_venta.all() if x.activo), None)
    return h.estado if h else 'venta_iniciada'


def _vendedor_nombre_cliente(cliente):
    from apps.persona.models import Vendedor
    r = cliente.respuestas_formulario.filter(nombre_campo__iexact='vendedor').first()
    if not r or not r.respuesta_campo:
        return None
    try:
        v = Vendedor.objects.filter(fecha_elimina__isnull=True).get(id=int(r.respuesta_campo))
        return v.nombre_completo
    except (ValueError, Vendedor.DoesNotExist):
        return r.respuesta_campo


def _empresa_servicio_producto_para_cliente(cliente):
    """
    Devuelve empresa_nombre, servicio_nombre, producto para un cliente (primera relación
    ClienteEmpresa o fallback desde cliente.servicio_id / cliente.producto). Útil para Excel.
    """
    from apps.servicio.models import Servicio
    # Prefetch ya trae cliente_empresas con estado='1' y select_related
    ce = next(iter(cliente.cliente_empresas.all()), None)
    if ce:
        return (
            (ce.empresa.nombre if ce.empresa else ''),
            (ce.servicio.nombre if ce.servicio else ''),
            ce.producto or '',
        )
    servicio = Servicio.objects.filter(id=cliente.servicio_id).select_related('empresa').first() if cliente.servicio_id else None
    empresa_nombre = servicio.empresa.nombre if servicio and servicio.empresa_id else ''
    servicio_nombre = servicio.nombre if servicio else ''
    producto = (cliente.producto or '').strip()
    return (empresa_nombre, servicio_nombre, producto)


def _formatear_estado_venta_legible(valor):
    """Estado de venta en formato legible para Excel (ej. Venta Iniciada, Completada)."""
    if not valor or not str(valor).strip():
        return ''
    v = str(valor).strip().lower().replace('-', ' ')
    mapeo = {
        'venta_iniciada': 'Venta Iniciada',
        'venta iniciada': 'Venta Iniciada',
        'completada': 'Completada',
        'pendiente': 'Pendiente',
        'cancelada': 'Cancelada',
        'en_proceso': 'En Proceso',
        'en proceso': 'En Proceso',
    }
    if v in mapeo:
        return mapeo[v]
    return ' '.join(p.capitalize() for p in v.split('_') if p)


def _productos_para_pdf(cliente):
    """
    Devuelve una lista de diccionarios con la información de cada producto del cliente
    para generar una página del PDF por cada uno. Si no hay ClienteEmpresa, se arma
    un producto a partir de cliente.servicio_id y cliente.producto.
    """
    from apps.servicio.models import Servicio
    # Usar prefetch (get_queryset ya incluye ClienteEmpresa con estado='1' y select_related)
    empresas = list(cliente.cliente_empresas.all())
    if empresas:
        return [
            {
                'empresa_nombre': (ce.empresa.nombre if ce.empresa else '-'),
                'servicio_nombre': (ce.servicio.nombre if ce.servicio else '-'),
                'producto': ce.producto or '-',
                'tipo_cliente': ce.tipo_cliente or '-',
                'vendedor': _vendedor_por_producto(cliente, ce),
            }
            for ce in empresas
        ]
    # Cliente sin ClienteEmpresa (registro antiguo): un solo “producto” con servicio/producto del cliente
    servicio = Servicio.objects.filter(id=cliente.servicio_id).select_related('empresa').first() if cliente.servicio_id else None
    empresa_nombre = servicio.empresa.nombre if servicio and servicio.empresa_id else '-'
    servicio_nombre = servicio.nombre if servicio else '-'
    producto = (cliente.producto or '').strip() or '-'
    vendedor = _vendedor_nombre_cliente(cliente) or '-'
    return [{'empresa_nombre': empresa_nombre, 'servicio_nombre': servicio_nombre, 'producto': producto, 'tipo_cliente': '-', 'vendedor': vendedor}]


def _estado_venta_por_producto(cliente, cliente_empresa):
    """Estado de venta del producto (ClienteEmpresa). Si cliente_empresa es None, usa estado a nivel cliente."""
    if cliente_empresa:
        # Historial de ESTE producto (prefetched)
        for h in cliente_empresa.historial_estados_venta.all():
            if h and h.activo:
                return h.estado or 'venta_iniciada'
        # Legacy: historial con cliente_empresa null (prefetched en cliente)
        h_legacy = next((x for x in cliente.historial_estados_venta.all() if x.cliente_empresa_id is None and x.activo), None)
        return h_legacy.estado if h_legacy else 'venta_iniciada'
    return _estado_venta_cliente(cliente)


def _nombre_persona(p):
    """Nombre legible de una Persona (usuario_registra)."""
    if not p:
        return ''
    return getattr(p, 'nombre_completo', None) or (f'{getattr(p, "first_name", "")} {getattr(p, "last_name", "")}'.strip()) or str(p)


def _vendedor_por_producto(cliente, cliente_empresa):
    """
    Vendedor del producto. La fuente principal es ClienteEmpresa.vendedor (tabla cliente_empresa).
    Fallbacks: historial del producto, legacy, usuario_registra del ce, formulario cliente.
    """
    if cliente_empresa:
        # 1) Vendedor del producto en ClienteEmpresa (relación directa producto-vendedor)
        if cliente_empresa.vendedor_id:
            v = cliente_empresa.vendedor
            return getattr(v, 'nombre_completo', None) or str(v) if v else ''
        # 2) Historial activo de este producto -> usuario_registra
        for h in cliente_empresa.historial_estados_venta.all():
            if h and h.usuario_registra_id:
                return _nombre_persona(h.usuario_registra)
        # 3) Historial legacy (cliente_empresa null)
        h_legacy = next((x for x in cliente.historial_estados_venta.all() if x.cliente_empresa_id is None and x.activo), None)
        if h_legacy and h_legacy.usuario_registra_id:
            return _nombre_persona(h_legacy.usuario_registra)
        # 4) Quien registró este producto (al agregar producto)
        if cliente_empresa.usuario_registra_id:
            return _nombre_persona(cliente_empresa.usuario_registra)
    return _vendedor_nombre_cliente(cliente) or ''


def _productos_para_excel(cliente):
    """
    Lista de dicts por producto del cliente para Excel. Estado y vendedor son POR PRODUCTO:
    del HistorialEstadoVenta de ese producto (cliente_empresa), no del último ni del cliente.
    """
    from apps.servicio.models import Servicio
    empresas = list(cliente.cliente_empresas.all())
    if empresas:
        return [
            {
                'empresa_nombre': (ce.empresa.nombre if ce.empresa else ''),
                'servicio_nombre': (ce.servicio.nombre if ce.servicio else ''),
                'producto': ce.producto or '',
                'estado_venta': _estado_venta_por_producto(cliente, ce),
                'vendedor': _vendedor_por_producto(cliente, ce),
            }
            for ce in empresas
        ]
    servicio = Servicio.objects.filter(id=cliente.servicio_id).select_related('empresa').first() if cliente.servicio_id else None
    empresa_nombre = servicio.empresa.nombre if servicio and servicio.empresa_id else ''
    servicio_nombre = servicio.nombre if servicio else ''
    producto = (cliente.producto or '').strip()
    estado_venta = _estado_venta_cliente(cliente)
    vendedor = _vendedor_nombre_cliente(cliente) or ''
    return [{
        'empresa_nombre': empresa_nombre,
        'servicio_nombre': servicio_nombre,
        'producto': producto,
        'estado_venta': estado_venta,
        'vendedor': vendedor,
    }]


def _estilo_tabla_base():
    """Estilo común para tablas del PDF (bordes, padding, fuente)."""
    return [
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
    ]


def _formatear_estado_venta(valor):
    """Convierte valor interno de estado de venta a texto en MAYÚSCULAS (ej. venta_iniciada → VENTA INICIADA)."""
    if not valor or not str(valor).strip():
        return '-'
    v = str(valor).strip().lower().replace('-', ' ')
    # Mapeos conocidos en mayúsculas para consistencia con el resto del PDF
    mapeo = {
        'venta_iniciada': 'VENTA INICIADA',
        'venta iniciada': 'VENTA INICIADA',
        'completada': 'COMPLETADA',
        'pendiente': 'PENDIENTE',
        'cancelada': 'CANCELADA',
        'en_proceso': 'EN PROCESO',
        'en proceso': 'EN PROCESO',
    }
    if v in mapeo:
        return mapeo[v]
    # Genérico: mayúsculas
    return ' '.join(p.upper() for p in v.split('_') if p)


def _formatear_valor_campo(nombre_campo, valor, estado_venta_formateado=None, vendedor_nombre=None):
    """
    Formatea el valor de un campo para mostrarlo en el PDF.
    - Estado de venta: usa estado_venta_formateado si se pasa (ej. "Venta Iniciada").
    - Cambio de titular: 1, si, true → Sí; 0, no → No.
    - Vendedor: si vendedor_nombre está dado, se muestra el nombre en lugar del ID.
    """
    if valor is None:
        return '-'
    raw = str(valor).strip()
    norm_nombre = (nombre_campo or '').lower().replace(' ', '_').replace('-', ' ')
    if not raw and norm_nombre not in ('estado_venta', 'estado de venta', 'vendedor'):
        return '-'
    # Estado de venta: usar el texto ya formateado (ej. Venta Iniciada)
    if 'estado' in norm_nombre and 'venta' in norm_nombre:
        if estado_venta_formateado is not None:
            return estado_venta_formateado
        return _formatear_estado_venta(raw)
    # Vendedor: mostrar nombre resuelto si está disponible
    if norm_nombre == 'vendedor':
        if vendedor_nombre:
            return vendedor_nombre
        return raw if raw else '-'
    # Cambio de titular: 1 = Sí, 0 = No
    if 'cambio' in norm_nombre and 'titular' in norm_nombre:
        if raw.lower() in ('1', 'si', 'sí', 'true', 'yes', 'verdadero'):
            return 'Sí'
        if raw.lower() in ('0', 'no', 'false', 'none', ''):
            return 'No'
        return raw
    if raw in ('1', '0'):
        return 'Sí' if raw == '1' else 'No'
    return raw


def _formatear_etiqueta_campo(nombre_campo):
    """Devuelve etiqueta legible en MAYÚSCULAS para consistencia (ej. tipo_cliente → TIPO DE CLIENTE)."""
    if not nombre_campo:
        return '-'
    s = str(nombre_campo).strip()
    if s.lower() in ('cambio de titular', 'cambio titular'):
        return 'CAMBIO DE TITULAR'
    if '_' in s and ' ' not in s:
        return s.replace('_', ' ').upper()
    return s.upper()


class ClienteViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_class = ClienteFilter
    search_fields = ['nombre', 'numero_identificacion', 'correo_electronico_o_carta', 'telefono']

    def get_queryset(self):
        qs = Cliente.objects.filter(estado='1').order_by('-fecha_registro')
        prefetch = [
            'respuestas_formulario',
            Prefetch(
                'historial_estados_venta',
                queryset=HistorialEstadoVenta.objects.select_related('usuario_registra'),
            ),
            Prefetch(
                'cliente_empresas',
                queryset=ClienteEmpresa.objects.filter(estado='1')
                .select_related('empresa', 'servicio', 'usuario_registra', 'vendedor', 'cerrador')
                .prefetch_related(
                    Prefetch(
                        'historial_estados_venta',
                        queryset=HistorialEstadoVenta.objects.filter(activo=True).select_related('usuario_registra'),
                    )
                )
                .order_by('id'),
            ),
        ]
        qs = qs.prefetch_related(*prefetch)
        return qs

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ClienteDetalleSerializer
        if self.action in ('update', 'partial_update'):
            return ClienteUpdateSerializer
        return ClienteSerializer

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.estado = '0'
        instance.usuario_elimina = request.user
        instance.fecha_elimina = timezone.now()
        instance.save()
        return Response(
            {'mensaje': 'Cliente eliminado correctamente.'},
            status=status.HTTP_200_OK,
        )

    def create(self, request, *args, **kwargs):
        serializer = ClienteCreateSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        cliente = serializer.save()
        return Response(
            {
                'mensaje': 'Cliente creado exitosamente.',
                'data': ClienteSerializer(cliente).data,
            },
            status=status.HTTP_201_CREATED,
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = ClienteUpdateSerializer(instance, data=request.data, partial=partial, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({
            'mensaje': 'Cliente actualizado correctamente.',
            'data': ClienteDetalleSerializer(instance).data,
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='agregar-producto')
    def agregar_producto(self, request, pk=None):
        """Agrega un nuevo producto a un cliente existente (sin duplicar datos del cliente)."""
        cliente = self.get_object()
        serializer = ClienteAgregarProductoSerializer(
            data=request.data,
            context={'request': request, 'cliente': cliente}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(
            {'mensaje': 'Producto agregado correctamente.'},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['post'], url_path='actualizar-producto')
    def actualizar_producto(self, request, pk=None):
        """Actualiza un producto (ClienteEmpresa) existente, incluido el vendedor del producto."""
        cliente = self.get_object()
        serializer = ClienteActualizarProductoSerializer(
            data=request.data,
            context={'request': request, 'cliente': cliente},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(
            {'mensaje': 'Producto actualizado correctamente.'},
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=['post'], url_path='cambiar-estado')
    def cambiar_estado(self, request, pk=None):
        """
        Cambia el estado de venta (y opcionalmente el vendedor) de un producto o del cliente.
        - cliente_empresa_id: si se envía, actualiza el historial de ESE producto; si no, a nivel cliente (legacy).
        - vendedor_id: opcional; id de Persona que quedará como usuario_registra en el historial.
        """
        cliente = self.get_object()
        nuevo_estado = (request.data.get('estado') or '').strip() or 'pendiente'
        cliente_empresa_id = request.data.get('cliente_empresa_id')
        vendedor_id = request.data.get('vendedor_id')
        cliente_empresa = None
        if cliente_empresa_id is not None:
            try:
                cliente_empresa = ClienteEmpresa.objects.get(
                    pk=cliente_empresa_id,
                    cliente=cliente,
                    estado='1',
                )
            except (ClienteEmpresa.DoesNotExist, ValueError, TypeError):
                return Response(
                    {'error': 'cliente_empresa_id no válido o no pertenece a este cliente.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        usuario_registra = None
        if vendedor_id is not None:
            from apps.persona.models import Persona
            try:
                usuario_registra = Persona.objects.get(pk=vendedor_id)
            except (Persona.DoesNotExist, ValueError, TypeError):
                pass
        _cambiar_estado_venta(
            cliente, nuevo_estado, request.user,
            cliente_empresa=cliente_empresa,
            usuario_registra=usuario_registra,
        )
        return Response({
            'mensaje': 'Estado actualizado correctamente.',
            'estado': nuevo_estado,
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'], url_path='descargar-pdf')
    def descargar_pdf(self, request, pk=None):
        """
        Genera un PDF con una página por cada producto del cliente.
        Cada página incluye: datos del cliente, información del producto (empresa/servicio/producto)
        y respuestas del formulario, con espacio reservado para contrato y firmas.
        """
        cliente = self.get_object()
        productos = _productos_para_pdf(cliente)
        respuestas = sorted(
            [r for r in cliente.respuestas_formulario.all() if r.estado == '1'],
            key=lambda r: (r.nombre_campo or ''),
        )
        estado_venta_raw = _estado_venta_cliente(cliente)
        estado_venta = _formatear_estado_venta(estado_venta_raw) if estado_venta_raw else '-'
        fecha_generacion = timezone.now().strftime('%d/%m/%Y %H:%M')

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.8 * cm,
            leftMargin=1.8 * cm,
            topMargin=1.8 * cm,
            bottomMargin=2.2 * cm,
        )
        styles = getSampleStyleSheet()
        color_titulo = colors.HexColor('#1e3a5f')
        color_seccion = colors.HexColor('#2c5282')
        color_etiqueta = colors.HexColor('#4a5568')
        color_valor = colors.HexColor('#2d3748')
        color_bg_etiqueta = colors.HexColor('#edf2f7')
        color_pie = colors.HexColor('#718096')

        def _pie_pagina(canvas, document):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(color_pie)
            w = document.pagesize[0]
            canvas.drawCentredString(w / 2, 1.2 * cm, f'Documento generado el {fecha_generacion}')
            canvas.restoreState()

        header_style = ParagraphStyle(
            'pdf_header',
            parent=styles['Normal'],
            fontSize=8,
            textColor=color_pie,
            spaceAfter=2,
            alignment=2,
        )
        title_style = ParagraphStyle(
            'pdf_title',
            parent=styles['Heading1'],
            fontSize=13,
            spaceBefore=0,
            spaceAfter=10,
            textColor=color_titulo,
            fontName='Helvetica-Bold',
        )
        section_style = ParagraphStyle(
            'pdf_section',
            parent=styles['Heading2'],
            fontSize=10,
            spaceBefore=16,
            spaceAfter=6,
            textColor=color_seccion,
            fontName='Helvetica-Bold',
        )
        note_style = ParagraphStyle(
            'pdf_note',
            parent=styles['Normal'],
            fontSize=9,
            textColor=color_pie,
        )
        elements = []

        norm_campo = lambda s: (s or '').lower().strip().replace(' ', '_')
        respuestas_sin_vendedor = [r for r in respuestas if norm_campo(r.nombre_campo) != 'vendedor']

        for idx, prod in enumerate(productos):
            if idx > 0:
                elements.append(PageBreak())

            elements.append(Paragraph('DOCUMENTO CONFIDENCIAL', header_style))
            elements.append(Spacer(1, 4))

            titulo_producto = prod['producto'] if prod['producto'] != '-' else f'Producto {idx + 1}'
            elements.append(Paragraph(f'ESPACIO DE CONTRATO — {titulo_producto}', title_style))
            elements.append(Spacer(1, 14))

            # 1. DATOS DEL CLIENTE (sin vendedor)
            elements.append(Paragraph('1. DATOS DEL CLIENTE', section_style))
            datos_cliente = [
                ['NOMBRE', cliente.nombre or '-'],
                ['TIPO DE IDENTIFICACIÓN', cliente.tipo_identificacion or '-'],
                ['NÚMERO DE IDENTIFICACIÓN', cliente.numero_identificacion or '-'],
                ['TELÉFONO', cliente.telefono or '-'],
                ['CORREO O CARTA', cliente.correo_electronico_o_carta or '-'],
                ['ESTADO DE VENTA', estado_venta],
            ]
            t_cli = Table(datos_cliente, colWidths=[5 * cm, 10.5 * cm])
            t_cli.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), color_bg_etiqueta),
                ('TEXTCOLOR', (0, 0), (0, -1), color_etiqueta),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('TEXTCOLOR', (1, 0), (1, -1), color_valor),
                * _estilo_tabla_base(),
            ]))
            elements.append(t_cli)

            # 2. EMPRESA Y PRODUCTO (relación cliente_empresas)
            elements.append(Paragraph('2. EMPRESA Y PRODUCTO ASOCIADO', section_style))
            datos_empresa = [
                ['SERVICIO', prod['empresa_nombre']],
                ['CONTRATISTA', prod['servicio_nombre']],
                ['PRODUCTO', prod['producto']],
                ['TIPO DE CLIENTE', prod['tipo_cliente']],
            ]
            t_emp = Table(datos_empresa, colWidths=[5 * cm, 10.5 * cm])
            t_emp.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), color_bg_etiqueta),
                ('TEXTCOLOR', (0, 0), (0, -1), color_etiqueta),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('TEXTCOLOR', (1, 0), (1, -1), color_valor),
                * _estilo_tabla_base(),
            ]))
            elements.append(t_emp)

            # 3. INFORMACIÓN DEL FORMULARIO (todos los campos excepto vendedor)
            elements.append(Paragraph('3. INFORMACIÓN DEL FORMULARIO', section_style))
            datos_form = [
                [
                    _formatear_etiqueta_campo(r.nombre_campo),
                    _formatear_valor_campo(
                        r.nombre_campo,
                        r.respuesta_campo,
                        estado_venta_formateado=estado_venta,
                        vendedor_nombre=None,
                    ),
                ]
                for r in respuestas_sin_vendedor
            ]
            if datos_form:
                t_form = Table(datos_form, colWidths=[5 * cm, 10.5 * cm])
                t_form.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), color_bg_etiqueta),
                    ('TEXTCOLOR', (0, 0), (0, -1), color_etiqueta),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('TEXTCOLOR', (1, 0), (1, -1), color_valor),
                    * _estilo_tabla_base(),
                ]))
                elements.append(t_form)
            else:
                elements.append(Paragraph('Sin respuestas de formulario registradas.', note_style))

            # 4. DATOS DEL VENDEDOR (vendedor por producto, como en Excel)
            vendedor_producto = prod.get('vendedor') or '-'
            elements.append(Paragraph('4. DATOS DEL VENDEDOR', section_style))
            datos_vendedor = [
                ['VENDEDOR ASIGNADO', vendedor_producto],
            ]
            t_vend = Table(datos_vendedor, colWidths=[5 * cm, 10.5 * cm])
            t_vend.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), color_bg_etiqueta),
                ('TEXTCOLOR', (0, 0), (0, -1), color_etiqueta),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('TEXTCOLOR', (1, 0), (1, -1), color_valor),
                * _estilo_tabla_base(),
            ]))
            elements.append(t_vend)

        doc.build(elements, onFirstPage=_pie_pagina, onLaterPages=_pie_pagina)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        nombre_safe = (cliente.nombre or 'contrato').replace(' ', '_')[:50]
        response['Content-Disposition'] = f'attachment; filename="cliente_{cliente.id}_{nombre_safe}.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta clientes a Excel: una fila por producto.
        Columnas: NOMBRE, TIPO IDENTIFICACIÓN, NÚMERO IDENTIFICACIÓN, TELÉFONO, CORREO,
        SERVICIO, COMPAÑÍA ACTUAL, TIPO PRODUCTO, ESTADO VENTA, VENDEDOR.
        Estado de venta y vendedor por producto. Datos del cliente unificados con merge vertical.
        """
        queryset = self.filter_queryset(self.get_queryset())
        wb = Workbook()
        ws = wb.active
        ws.title = 'Clientes'

        def _mayus(s):
            return (s or '').strip().upper() if isinstance(s, str) else str(s or '').upper()

        # SERVICIO = antes "Tipo de empresa"; COMPAÑÍA ACTUAL = antes "Tipo de servicio"; orden: SERVICIO | COMPAÑÍA ACTUAL | TIPO PRODUCTO
        headers = [
            'NOMBRE', 'TIPO IDENTIFICACIÓN', 'NÚMERO IDENTIFICACIÓN', 'TELÉFONO', 'CORREO ELECTRÓNICO O CARTA',
            'SERVICIO', 'COMPAÑÍA ACTUAL', 'TIPO PRODUCTO',
            'ESTADO VENTA', 'VENDEDOR',
        ]
        ws.append(headers)

        # Rango de filas por cliente para merge (columnas 1-5)
        merge_ranges = []  # [(start_row, end_row), ...]
        data_row = 2

        for c in queryset:
            productos = _productos_para_excel(c)
            start_row = data_row
            for idx, prod in enumerate(productos):
                estado_venta = _formatear_estado_venta(prod['estado_venta']) if prod.get('estado_venta') else ''
                vendedor = prod.get('vendedor') or ''
                if idx == 0:
                    ws.append([
                        _mayus(c.nombre),
                        _mayus(c.tipo_identificacion),
                        _mayus(c.numero_identificacion),
                        _mayus(c.telefono),
                        _mayus(c.correo_electronico_o_carta),
                        _mayus(prod['empresa_nombre']),
                        _mayus(prod['servicio_nombre']),
                        _mayus(prod['producto']),
                        _mayus(estado_venta),
                        _mayus(vendedor),
                    ])
                else:
                    ws.append([
                        '', '', '', '', '',
                        _mayus(prod['empresa_nombre']),
                        _mayus(prod['servicio_nombre']),
                        _mayus(prod['producto']),
                        _mayus(estado_venta),
                        _mayus(vendedor),
                    ])
                data_row += 1
            if len(productos) > 1:
                merge_ranges.append((start_row, data_row - 1))

        # Unificación visual: merge vertical de celdas de datos del cliente (columnas 1-5)
        for start_row, end_row in merge_ranges:
            for col in range(1, 6):
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

        # Estilo profesional: encabezado en negrita y fondo
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        for col, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        column_widths = [22, 18, 20, 14, 28, 22, 22, 22, 16, 22]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws.freeze_panes = 'A2'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='descargar-plantilla')
    def descargar_plantilla(self, request):
        """
        Descarga plantilla Excel para importación masiva de clientes.
        Columnas: Nombre completo, Tipo identificación, Nº identificación, CUPS, Dirección,
        Teléfono, Correo, Compañía anterior, Compañía actual, Producto, Correo electrónico o carta.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Clientes'

        headers = [
            'Nombre completo',
            'Tipo identificación',
            'Número identificación',
            'Cuenta bancaria',
            'Dirección',
            'Teléfono',
            'Correo electrónico o carta',
            'Compañía anterior',
            'Compañía actual',
            'Servicio',
            'Compañía actual',
            'Producto',
        ]
        ws.append(headers)

        # Estilo encabezado
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        for col, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Fila de ejemplo
        ws.append([
            'Ejemplo: Juan Pérez',
            'DNI',
            '123456789',
            'ES12 3456 7890 1234 5678 90',
            'Calle Ejemplo 123',
            '600123456',
            'ejemplo@correo.com',
            'Empresa anterior S.L.',
            'Empresa actual S.A.',
            'Nombre del servicio (empresa)',
            'Nombre de la compañía actual (servicio)',
            'Producto A',
        ])

        column_widths = [22, 18, 20, 28, 28, 14, 28, 22, 22, 28, 28, 18]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_importar_clientes.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa clientes desde Excel.
        Columnas: Nombre, Tipo identificación, Nº identificación, CUPS, Cuenta bancaria,
        Dirección, Teléfono, Correo, Compañía anterior, Compañía actual, Producto,
        Correo electrónico o carta.
        """
        archivo = request.FILES.get('archivo') or request.FILES.get('file')

        if not archivo:
            return Response(
                {'error': 'Debe adjuntar un archivo Excel.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {'error': f'No se pudo leer el archivo Excel: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Columnas: 0=Nombre, 1=Tipo id, 2=Nº id, 3=Cuenta bancaria, 4=Dirección,
        # 6=Teléfono, 7=Correo, 8=Compañía anterior, 9=Compañía actual, 10=Servicio, 11=Contratista, 12=Producto
        from apps.empresa.models import Empresa
        from apps.servicio.models import Servicio as ServicioModel

        creados = 0
        errores = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            nombre = (row[0] or '').strip() if len(row) > 0 else ''
            if not nombre:
                errores.append(f'Fila {row_idx}: Nombre vacío.')
                continue

            tipo_id = (row[1] or '').strip() if len(row) > 1 else ''
            numero_id = (row[2] or '') if len(row) > 2 else ''
            cuenta_bancaria = (row[3] or '').strip() if len(row) > 3 else ''
            direccion = (row[4] or '').strip() if len(row) > 4 else ''
            telefono = (row[5] or '') if len(row) > 5 else ''
            correo_electronico_o_carta = (row[6] or '').strip() if len(row) > 6 else ''
            compania_ant = (row[7] or '').strip() if len(row) > 7 else ''
            compania_act = (row[8] or '').strip() if len(row) > 8 else ''
            # Plantilla nueva: columnas 9=Servicio, 10=Contratista, 11=Producto. Antigua: 9=Producto.
            tiene_columnas_servicio_contratista = len(row) > 10
            if tiene_columnas_servicio_contratista:
                servicio_nombre = (row[9] or '').strip() if len(row) > 9 else ''
                contratista_nombre = (row[10] or '').strip() if len(row) > 10 else ''
                producto = (row[11] or '').strip() if len(row) > 11 else ''
            else:
                servicio_nombre = ''
                contratista_nombre = ''
                producto = (row[9] or '').strip() if len(row) > 9 else ''

            servicio_id = None
            if tiene_columnas_servicio_contratista and (servicio_nombre or contratista_nombre):
                empresa = Empresa.objects.filter(
                    nombre__iexact=servicio_nombre, estado='1'
                ).first() if servicio_nombre else None
                if not empresa and servicio_nombre:
                    errores.append(f'Fila {row_idx}: El servicio "{servicio_nombre}" no existe.')
                    continue
                if contratista_nombre:
                    servicio = ServicioModel.objects.filter(
                        empresa=empresa, nombre__iexact=contratista_nombre, estado='1'
                    ).first() if empresa else None
                    if not servicio:
                        errores.append(
                            f'Fila {row_idx}: El contratista "{contratista_nombre}" no existe'
                            + (f' para el servicio "{servicio_nombre}".' if servicio_nombre else '.')
                        )
                        continue
                    servicio_id = servicio.id
                elif empresa:
                    primer_servicio = ServicioModel.objects.filter(empresa=empresa, estado='1').first()
                    if primer_servicio:
                        servicio_id = primer_servicio.id

            payload = {
                'servicio_id': servicio_id,
                'nombre': nombre,
                'tipo_identificacion': tipo_id,
                'numero_identificacion': numero_id,
                'telefono': telefono,
                'correo_electronico_o_carta': correo_electronico_o_carta,
                'direccion': direccion,
                'cuenta_bancaria': cuenta_bancaria,
                'compania_anterior': compania_ant,
                'compania_actual': compania_act,
                'producto': producto,
                'respuestas': [],
            }

            serializer = ClienteCreateSerializer(data=payload, context={'request': request})
            if serializer.is_valid():
                cliente = serializer.save()
                cliente.creado_por_carga_masiva = True
                cliente.save(update_fields=['creado_por_carga_masiva'])
                creados += 1
            else:
                err_msg = '; '.join(
                    f'{k}: {v[0]}' if isinstance(v, list) else f'{k}: {v}'
                    for k, v in serializer.errors.items()
                )
                errores.append(f'Fila {row_idx}: {err_msg}')

        wb.close()

        return Response({
            'mensaje': f'Se importaron {creados} cliente(s) correctamente.',
            'creados': creados,
            'errores': errores[:50],
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='subir-documentos')
    def subir_documentos(self, request, pk=None):
        """Sube documento_dni y/o documento_factura. Permite primera carga para cualquier cliente."""
        cliente = self.get_object()
        doc_dni = request.FILES.get('documento_dni')
        doc_factura = request.FILES.get('documento_factura')
        if not doc_dni and not doc_factura:
            return Response(
                {'error': 'Debe adjuntar al menos documento_dni o documento_factura.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if doc_dni:
            cliente.documento_dni = doc_dni
        if doc_factura:
            cliente.documento_factura = doc_factura
        update_f = []
        if doc_dni:
            update_f.append('documento_dni')
        if doc_factura:
            update_f.append('documento_factura')
        cliente.save(update_fields=update_f)
        return Response({'mensaje': 'Documentos subidos correctamente.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'], url_path='documento-dni')
    def documento_dni(self, request, pk=None):
        """Devuelve el PDF del DNI del cliente."""
        cliente = self.get_object()
        if not cliente.documento_dni:
            return Response({'error': 'No hay documento DNI.'}, status=status.HTTP_404_NOT_FOUND)
        from django.http import FileResponse
        return FileResponse(cliente.documento_dni.open('rb'), as_attachment=False, content_type='application/pdf')

    @action(detail=True, methods=['get'], url_path='documento-factura')
    def documento_factura(self, request, pk=None):
        """Devuelve el PDF de la factura del cliente."""
        cliente = self.get_object()
        if not cliente.documento_factura:
            return Response({'error': 'No hay documento de factura.'}, status=status.HTTP_404_NOT_FOUND)
        from django.http import FileResponse
        return FileResponse(cliente.documento_factura.open('rb'), as_attachment=False, content_type='application/pdf')

    @action(detail=True, methods=['get'], url_path='descargar-documento-dni')
    def descargar_documento_dni(self, request, pk=None):
        """Descarga el PDF del DNI del cliente."""
        cliente = self.get_object()
        if not cliente.documento_dni:
            return Response({'error': 'No hay documento DNI.'}, status=status.HTTP_404_NOT_FOUND)
        from django.http import FileResponse
        response = FileResponse(cliente.documento_dni.open('rb'), as_attachment=True, filename='dni_cliente.pdf')
        response['Content-Type'] = 'application/pdf'
        return response

    @action(detail=True, methods=['get'], url_path='descargar-documento-factura')
    def descargar_documento_factura(self, request, pk=None):
        """Descarga el PDF de la factura del cliente."""
        cliente = self.get_object()
        if not cliente.documento_factura:
            return Response({'error': 'No hay documento de factura.'}, status=status.HTTP_404_NOT_FOUND)
        from django.http import FileResponse
        response = FileResponse(cliente.documento_factura.open('rb'), as_attachment=True, filename='factura_cliente.pdf')
        response['Content-Type'] = 'application/pdf'
        return response
